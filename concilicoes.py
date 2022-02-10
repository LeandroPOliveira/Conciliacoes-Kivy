from kivy.clock import Clock
from kivy.config import Config
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.spinner import Spinner
from kivymd.app import MDApp
from kivymd.uix.datatables import MDDataTable
from kivy.lang.builder import Builder
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.metrics import dp
import os
import pandas as pd
import threading
import openpyxl
import numpy as np
from datetime import datetime, date
from kivy.utils import get_color_from_hex
# from kivy.core.window import Window
# Window.size = (1280, 720)
from dateutil.relativedelta import relativedelta
Config.set('graphics', 'resizable', '1')
Config.set('graphics', 'width', '1280')
Config.set('graphics', 'height', '720')
Config.write()


class LoginWindow(Screen):
    pass


class DataWindow(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.lista = []
        self.resultado = None
        for i in range(12):
            mes = datetime.today()
            data_limite = mes - relativedelta(months=i)
            self.lista.append(data_limite.strftime('%m.%Y'))

    def on_spinner_select(self, text):
        self.text = text
        print(self.text)

    # def start_foo_thread(self, processo):
    #     # global foo_thread
    #     self.foo_thread = threading.Thread(target=processo)
    #     self.foo_thread.daemon = True
    #     self.popup = Popup(title='Test popup',
    #                   content=Label(text='Gerando relatório'),
    #                   size_hint=(None, None), size=(400, 400), auto_dismiss=False)
    #     self.popup.open()
    #
    #     self.foo_thread.start()
    #     Clock.schedule_interval(self.check_foo_thread, 5)
    #
    # def check_foo_thread(self, dt):
    #     if self.foo_thread.is_alive():
    #         Clock.schedule_interval(self.check_foo_thread, 5)
    #     else:
    #         self.popup.dismiss()


    def validacao(self):
        # self.resultado = None
        # self.caminho = 'G:\GECOT\CONCILIAÇÕES CONTÁBEIS\CONCILIAÇÕES_' + self.text[3:7] + \
        #                '\\' + self.text
        # pasta1 = os.listdir(self.caminho)
        #
        # lista = [[], [], [], [], []]
        #
        # for i in pasta1[::-1]:
        #     if i.startswith('~') == True:
        #         pasta1.remove(i)
        #
        #
        # for i in pasta1:
        #     if i.endswith('.xlsx'):
        #         wb = openpyxl.load_workbook(self.caminho + '\\' + i, read_only=True)
        #         sheets = wb.sheetnames
        #         ws = wb[sheets[0]]
        #         try:
        #             conta = ws['A2'].value.split()
        #         except:
        #             conta = ['', '']
        #         valor_deb = ws['C5'].value
        #         valor_cred = ws['D5'].value
        #         try:
        #             data1 = ws['A5'].value.strftime('%m.%Y')
        #         except:
        #             data1 = ws['A5'].value
        #         lista[0].append(conta[1])
        #         lista[1].append(data1)
        #         lista[2].append(valor_deb)
        #         lista[3].append(valor_cred)
        #
        #         wb.close()
        #
        # data = pd.DataFrame(lista).T
        #
        # data.columns = ['Conta', 'Data', 'Debito', 'Credito', 'Balancete']
        #
        # # Listar planilhas dos balancetes
        # try:
        #     self.pasta = 'G:\GECOT\CONCILIAÇÕES CONTÁBEIS\CONCILIAÇÕES_' + self.text[3:7] + \
        #             '\BALANCETES\SOCIETÁRIOS\\'
        #
        #     lista = os.listdir(self.pasta)
        #
        #     lista4 = {}
        #     for i in lista:
        #         if self.text in i or self.text.replace('.', '') in i:
        #             tempo = os.path.getmtime(self.pasta + i)
        #             tempo2 = datetime.fromtimestamp(tempo)
        #             lista4.update({i: tempo2})
        #
        #     dados = list(lista4.keys())[list(lista4.values()).index(max(lista4.values()))]
        #     dados = pd.read_excel(self.pasta + dados, skiprows=12)
        #
        # except:
        #     pass
        #     # tkinter.messagebox.showinfo('', 'Balancete não encontrado.\n Informar local do arquivo.')
        #     # dados = fd.askopenfilename(title='Abrir arquivo', initialdir=self.pasta)
        #     # dados = pd.read_excel(dados, skiprows=12)
        #
        #
        #
        # dados = pd.DataFrame(dados)
        #
        # apoio = pd.read_excel('contas.xlsx')
        # apoio = pd.DataFrame(apoio)
        #
        # for index1, row1 in data.iterrows():
        #     for index, row in dados.iterrows():
        #         if row1['Conta'] == row['Conta CSPE']:
        #             # data.insert(3, 'Saldo', '')
        #             data['Balancete'].loc[index1] = dados.loc[index, ' Saldo Acumulado']
        #
        # data[['Debito', 'Credito', 'Balancete']] = data[['Debito', 'Credito', 'Balancete']].apply(pd.to_numeric)
        #
        # data.fillna(0, inplace=True)
        # data = data.round(2)
        # data['Conciliação'] = data['Debito'] - data['Credito']
        #
        # data.drop(['Debito', 'Credito'], axis=1, inplace=True)
        #
        # data['Diferença'] = data['Conciliação'] - data['Balancete']
        #
        # # data['Resultado'] = data.apply(lambda x: x['Debito'] - x['Saldo'], axis=1)
        # data = pd.merge(data, apoio[['Conta', 'Usuario']], on=['Conta'], how='left')
        #
        # data['Status'] = np.where(data['Diferença'] != 0, 'Diferença de Valor',
        #                           (np.where(data['Data'] != self.text, 'Data Incorreta', 'OK')))
        #
        # #if self.usuario.get() != self.lista_usuarios[3]:
        #    # data = data.loc[data['Usuario'] == self.usuario.get()]
        #
        # data = data.round(2)
        # data.to_excel('teste.xlsx', index=False)

        dados = pd.read_excel('teste.xlsx')
        data = pd.DataFrame(dados)

        self.data = data.to_records(index=False)
        self.resultado = list(self.data)

        return self.resultado





class BoxTeste(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.data_tables = None

    def add_datatable(self):
        self.data_tables = MDDataTable(pos_hint={'center_x': 0.5, 'y': 0.2},
                                       size_hint=(0.8, 0.7),
                                       use_pagination=True, rows_num=10,
                                       background_color_header=get_color_from_hex("#03a9e0"),

                                       column_data=[("[color=#ffffff]Conta[/color]", dp(30)),
                                                    ("[color=#ffffff]Data[/color]", dp(20)),
                                                    ("[color=#ffffff]Balancete[/color]", dp(30)),
                                                    ("[color=#ffffff]Conciliação[/color]", dp(30)),
                                                    ("[color=#ffffff]Diferença[/color]", dp(20)),
                                                    ("[color=#ffffff]Usuario[/color]", dp(35)),
                                                    ("[color=#ffffff]Status[/color]", dp(35)),
                                                    ],
                                       row_data=DataWindow.validacao(self), elevation=1)

        self.add_widget(self.data_tables)




class WindowManager(ScreenManager):
    pass


class Example(MDApp):

    def build(self):


        return Builder.load_file('studentdb.kv')




    def change_screen(self, screen: str):
        self.root.current = screen


Example().run()