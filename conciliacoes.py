import sqlite3
import threading
from datetime import datetime
from dateutil.relativedelta import relativedelta
from kivy.clock import Clock
from kivy.core.window import Window
from kivy.properties import StringProperty, ListProperty, BooleanProperty
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivymd.uix.dialog import MDDialog
from kivymd.uix.menu import MDDropdownMenu
from kivy.utils import get_color_from_hex
import getpass
from kivymd.app import MDApp
from kivymd.uix.datatables import MDDataTable
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.metrics import dp
import numpy as np
import os
from os.path import basename
import pandas as pd
from PyPDF2 import PdfFileWriter, PdfFileReader
import openpyxl
from reportlab.pdfgen import canvas
from win32com import client
from zipfile import ZipFile


class TelaLogin(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.menu = None
        self.dialog = None
        self.nome_usuario = {}
        self.lista_usuarios = []
        with open('usuarios.txt') as users:
            self.usuarios = users.readlines()
            for user in self.usuarios:
                self.lista_usuarios.append(user.split(';')[0])
                self.nome_usuario.update({user.split(';')[0]: user.split(';')[1]})

    def abre_menu(self):
        menu_items = [{"viewclass": "OneLineListItem", "text": i, "height": dp(56), "on_release":
                      lambda x=i: self.set_item(x), } for i in self.lista_usuarios]

        self.menu = MDDropdownMenu(caller=self.ids.drop_item, items=menu_items, position="center", width_mult=4, )
        self.menu.open()

    def set_item(self, text_item):
        self.ids.drop_item.text = text_item
        self.menu.dismiss()

    def verifica_usuario(self):
        if self.nome_usuario[self.ids.drop_item.text].strip() == getpass.getuser():
            self.manager.current = 'validar'
        else:
            self.dialog = MDDialog(text="Usuário inválido!", radius=[20, 7, 20, 7], )
            self.dialog.open()


class TelaValidacao(Screen):
    meu_status = StringProperty('Não Verificado')
    meu_status1 = StringProperty('Não Verificado')
    meu_status2 = StringProperty('Não Verificado')
    status_btn = BooleanProperty(False)
    icone = StringProperty("alert")
    icone1 = StringProperty("alert")
    icone2 = StringProperty("alert")
    cor, cor1, cor2 = ListProperty([1, 1, 0, 1]), ListProperty([1, 1, 0, 1]), ListProperty([1, 1, 0, 1])

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.caminho_mes = None
        self.dialog = None
        self.intervalo_meses = []
        self.validos = []

        for i in range(24):
            mes = datetime.today()
            data_limite = mes - relativedelta(months=i)
            self.intervalo_meses.append(data_limite.strftime('%m.%Y'))

    def mes_selecionado(self, text):
        self.mes_atual = text
        # Arquivo que serve como base para identificar a pasta de trabalho a ser utilizada, o ano é alterado de acordo
        # com a competencia selecionada pelo usuário
        with open('pasta.txt', 'r', encoding='utf-8') as path:
            self.dados = path.readlines()

            self.caminho = os.path.join(self.dados[0].strip(), getpass.getuser(),
                                        self.dados[1][0:85] + '_' + self.mes_atual[3:9], self.dados[1][93:].rstrip())

            self.caminho_balancete = os.path.join(self.caminho, 'BALANCETES\SOCIETÁRIOS')


        # with open('pasta.txt', 'w', encoding='utf-8') as path:
        #     path.write(self.dados[0])
        #     path.write(f'{self.caminho[]}\n')

        return self.mes_atual, self.caminho

    def status(self):  # Verifica se a situação das conciliações, se está validada ou pendente
        self.validos.clear()
        with open('dados.txt', 'r') as f:
            lines = f.readlines()[1:]
            self.meu_status, self.meu_status1, self.meu_status2 = '', '', ''
            for i in lines:  # verifica a situação da competência selecionada, usuarios que validaram ou não
                i = i.split(';')
                if i[0] == self.mes_atual and i[1] == self.manager.get_screen('login').lista_usuarios[1] \
                        and i[2].strip() == 'OK':
                    self.meu_status, self.icone, self.cor = 'Validado', 'check-circle', [0, 1, 0, 1]
                    self.validos.append('OK')
                elif i[0] == self.mes_atual and i[1] == self.manager.get_screen('login').lista_usuarios[2] \
                        and i[2].strip() == 'OK':
                    self.meu_status1, self.icone1, self.cor1 = 'Validado', 'check-circle', [0, 1, 0, 1]
                    self.validos.append('OK')
                elif i[0] == self.mes_atual and i[1] == self.manager.get_screen('login').lista_usuarios[3] \
                        and i[2].strip() == 'OK':
                    self.meu_status2, self.icone2, self.cor2 = 'Validado', 'check-circle', [0, 1, 0, 1]
                    self.validos.append('OK')
                else:
                    if self.meu_status == '':
                        self.meu_status, self.icone, self.cor = 'Validação Pendente', 'alert-circle', [1, 0, 0, 1]
                    if self.meu_status1 == '':
                        self.meu_status1, self.icone1, self.cor1 = 'Validação Pendente', 'alert-circle', [1, 0, 0, 1]
                    if self.meu_status2 == '':
                        self.meu_status2, self.icone2, self.cor2 = 'Validação Pendente', 'alert-circle', [1, 0, 0, 1]

        usuario = self.manager.get_screen('login').ids.drop_item.text
        if usuario == self.manager.get_screen('login').lista_usuarios[0] and len(self.validos) >= 3:
            self.status_btn = True
        else:
            self.status_btn = False

    def start_foo_thread(self):
        self.foo_thread = threading.Thread(target=self.assina_gestor)
        self.foo_thread.daemon = True
        self.pb = MDDialog(text="Aguarde...", radius=[20, 7, 20, 7], )
        self.pb.open()
        self.foo_thread.start()
        Clock.schedule_interval(self.check_foo_thread, 10)

    def check_foo_thread(self, dt):
        if self.foo_thread.is_alive():
            Clock.schedule_interval(self.check_foo_thread, 10)
        else:
            self.pb.dismiss()
            self.dialog = MDDialog(text="Assinado com sucesso!", radius=[20, 7, 20, 7], )
            self.dialog.open()
            Clock.unschedule(self.check_foo_thread)

    def assina_gestor(self):
        self.caminho_mes = os.path.join(self.caminho, self.ids.spinner_id2.text)
        # Criar pdf com assinatura
        c = canvas.Canvas('watermark.pdf')
        # posicionar a imagem da assinatura nas coordenadas x e y
        c.drawImage(self.manager.get_screen('login').ids.drop_item.text + '.png', 350, 40, 150, 90,
                    mask='auto')
        c.save()
        watermark = PdfFileReader(open("watermark.pdf", "rb"))

        for file in os.listdir(self.caminho_mes):
            if file.endswith(".pdf"):
                output_file = PdfFileWriter()
                with open(self.caminho_mes + '\\' + file, "rb") as f:
                    input_file = PdfFileReader(f)
                    # Number of pages in input document
                    page_count = input_file.getNumPages()

                    # Go through all the input file pages to add a watermark to them
                    for page_number in range(page_count):
                        input_page = input_file.getPage(page_number)
                        if page_number == page_count - 1:
                            input_page.mergePage(watermark.getPage(0))
                        output_file.addPage(input_page)

                    with open(self.caminho_mes + '\\' + file[8:], "wb") as outputStream:
                        output_file.write(outputStream)

                with ZipFile(os.path.join(self.caminho_mes, self.ids.spinner_id2.text + '.zip'), 'a') as zip_arquivo:
                    zip_arquivo.write(os.path.join(self.caminho_mes, file[8:]),
                                      basename(os.path.join(self.caminho_mes, file[8:])))

                os.remove(self.caminho_mes + '\\' + file)
                os.remove(self.caminho_mes + '\\' + file[8:])


class TelaRelatorio(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.dialog_erro = None
        self.dialog_ok = None
        self.resultado = None
        self.data = None
        self.pasta_balancetes = None
        self.caminho_mes = None
        self.tabela_dados = None
        self.lista_usuarios = []
        with open('usuarios.txt') as user:
            usuarios = user.readlines()
            for u in usuarios:
                self.lista_usuarios.append(u.split(';')[0].strip())

    def validacao(self):
        self.caminho_mes = os.path.join(self.manager.get_screen('validar').caminho,
                                        self.manager.get_screen('validar').ids.spinner_id2.text)

        conciliacoes_do_mes = os.listdir(self.caminho_mes)
        relatorio = [[], [], [], [], []]
        for i in conciliacoes_do_mes[::-1]:
            if i.startswith('~'):
                conciliacoes_do_mes.remove(i)

        for conciliacao in conciliacoes_do_mes:
            if conciliacao.startswith('Conta') or conciliacao.startswith('conta'):

                wb = openpyxl.load_workbook(os.path.join(self.caminho_mes, conciliacao),
                                            read_only=True)
                sheets = wb.sheetnames
                ws = wb[sheets[0]]
                try:
                    conta = ws['A2'].value.split()
                except AttributeError:
                    conta = ['', '']
                valor_deb = ws['C5'].value
                valor_cred = ws['D5'].value
                try:
                    data1 = ws['A5'].value.strftime('%m.%Y')
                except AttributeError:
                    data1 = ws['A5'].value
                relatorio[0].append(conta[1])
                relatorio[1].append(data1)
                relatorio[2].append(valor_deb)
                relatorio[3].append(valor_cred)
                wb.close()

        self.data = pd.DataFrame(relatorio).T
        self.data.columns = ['Conta', 'Data', 'Debito', 'Credito', 'Balancete']
        # Listar planilhas dos balancetes
        self.pasta_balancetes = self.manager.get_screen('validar').caminho_balancete

        lista_balancetes = os.listdir(self.pasta_balancetes)

        for balancete in lista_balancetes:
            if self.manager.get_screen('validar').ids.spinner_id2.text in balancete or \
                    self.manager.get_screen('validar').ids.spinner_id2.text.replace('.', '') in balancete:
                balancete_atual = balancete

                dados = pd.read_excel(os.path.join(self.pasta_balancetes, balancete_atual), sheet_name=0)
                dados = pd.DataFrame(dados)
                conn = sqlite3.connect('contas')
                cursor = conn.cursor()
                cursor.execute('select * from cadastro order by Conta')
                tabela_contas = cursor.fetchall()
                tabela_contas = pd.DataFrame(tabela_contas)
                tabela_contas.columns = ['Conta', 'Usuario']
                for index1, row1 in self.data.iterrows():
                    for index, row in dados.iterrows():
                        if row1['Conta'] == row['Conta CSPE']:
                            self.data['Balancete'].loc[index1] = dados.loc[index, 'Saldo Acumulado']
                self.data[['Debito', 'Credito', 'Balancete']] = self.data[['Debito', 'Credito', 'Balancete']].apply(
                    pd.to_numeric, errors='coerce')
                self.data.fillna(0, inplace=True)
                self.data = self.data.round(2)
                self.data['Conciliação'] = self.data['Debito'] - self.data['Credito']
                self.data.drop(['Debito', 'Credito'], axis=1, inplace=True)
                self.data['Diferença'] = self.data['Conciliação'] - self.data['Balancete']
                self.data = pd.merge(self.data, tabela_contas[['Conta', 'Usuario']], on=['Conta'], how='left')
                self.data['Status'] = np.where(self.data['Diferença'] != 0, 'Diferença de Valor',
                                               (np.where(
                                                   self.data['Data'] != self.manager.get_screen('validar').ids.spinner_id2.text,
                                                   'Data Incorreta', 'OK')))
                if self.manager.get_screen('login').ids.drop_item.text != self.lista_usuarios[0]:
                    self.data = self.data.loc[self.data['Usuario'] == self.manager.get_screen('login').ids.drop_item.text]
                self.data = self.data.round(2)
                self.resultado = self.data.to_records(index=False)
                self.resultado = list(self.resultado)
                if len(self.resultado) == 1 or isinstance(len(self.resultado) / 10, float):
                    self.resultado.append(('', '', '', '', '', '', ''))
                self.add_tabela()

    def add_tabela(self):
        self.tabela_dados = MDDataTable(pos_hint={'center_x': 0.5, 'y': 0.2},
                                        size_hint=(0.7, 0.7),
                                        use_pagination=True, rows_num=10,
                                        background_color_header=get_color_from_hex("#03a9e0"),
                                        check=True,
                                        column_data=[("[color=#ffffff]Conta[/color]", dp(40)),
                                                     ("[color=#ffffff]Data[/color]", dp(20)),
                                                     ("[color=#ffffff]Balancete[/color]", dp(30)),
                                                     ("[color=#ffffff]Conciliação[/color]", dp(30)),
                                                     ("[color=#ffffff]Diferença[/color]", dp(20)),
                                                     ("[color=#ffffff]Usuario[/color]", dp(35)),
                                                     ("[color=#ffffff]Status[/color]", dp(35)),
                                                     ],
                                        row_data=self.resultado, elevation=1)

        self.add_widget(self.tabela_dados)
        self.tabela_dados.bind(on_check_press=self.item_marcado)

    def item_marcado(self, instance_table, current_row):
        os.startfile(os.path.join(self.caminho_mes, 'Conta ' +
                                  current_row[0].replace('.', '') + '.xlsx'))

    def start_foo_thread2(self):
        valida = self.data['Status'].unique()
        if 'OK' in valida and len(valida) == 1:  # checa se não existem contas que não estão validadas
            self.foo_thread = threading.Thread(target=self.assina)
            self.foo_thread.daemon = True
            self.pb = MDDialog(text="Aguarde...", radius=[20, 7, 20, 7], )
            self.pb.open()
            self.foo_thread.start()
            Clock.schedule_interval(self.check_foo_thread2, 10)
        else:
            self.dialog_erro = MDDialog(text="Erro! Verificar pendências!", radius=[20, 7, 20, 7], )
            self.dialog_erro.open()

    def check_foo_thread2(self, dt):
        if self.foo_thread.is_alive():
            Clock.schedule_interval(self.check_foo_thread2, 10)
        else:
            self.pb.dismiss()
            self.dialog = MDDialog(text="Assinado com sucesso!", radius=[20, 7, 20, 7], )
            self.dialog.open()
            Clock.unschedule(self.check_foo_thread2)

    def assina(self):
        lista3 = []
        for i in self.data['Conta']:
            conta = 'Conta ' + i.replace('.', '') + '.xlsx'
            lista3.append(conta)

        # Criar pdf com assinatura
        c = canvas.Canvas('watermark.pdf')
        # posicionar a imagem da assinatura nas coordenadas x e y
        c.drawImage(self.manager.get_screen('login').ids.drop_item.text + '.png', 40, 50, 150, 100,
                    mask='auto')
        c.save()

        for i in lista3:  # converter arquivos excel para pdf
            excel = client.Dispatch("Excel.Application")
            sheets = excel.Workbooks.Open(self.caminho_mes + '\\' + i)
            work_sheets = sheets.Worksheets[1]
            path = os.path.join(self.caminho_mes, 'teste ' + i.replace('.xlsx', '.pdf'))
            work_sheets.ExportAsFixedFormat(0, path)

            # Buscar a assinatura criada
            watermark = PdfFileReader(open("watermark.pdf", "rb"))
            output = PdfFileWriter()

            with open(path, "rb") as provisorio:
                arquivo = PdfFileReader(provisorio)
                number_of_pages = arquivo.getNumPages()

                for current_page_number in range(number_of_pages):
                    page = arquivo.getPage(current_page_number)
                    if page.extractText() != "":
                        output.addPage(page)

                page_count = output.getNumPages()
                output2 = PdfFileWriter()
                # percorrer o arquivo para unir o pdf de assinatura na última pagina do arquivo principal
                for page_number in range(page_count):
                    input_page = output.getPage(page_number)
                    if page_number == page_count - 1:
                        input_page.mergePage(watermark.getPage(0))
                    output2.addPage(input_page)

                # exportar o arquivo pdf com o nome sufixo pendente, que aguardará assinatura do gestor
                with open(os.path.join(self.caminho_mes, 'pendente' +
                                                         i.replace('.xlsx', '.pdf')), "wb") as outputStream:
                    output2.write(outputStream)

            os.remove(path)
            sheets.Close(True)

        #  Após assinatura, criar um log simples em um arquivo de texto que permitirá ao sistema verificar o
        #  usuario que validou a sua respectiva conciliação e o período
        adicionar = [self.manager.get_screen('validar').ids.spinner_id2.text,
                     self.manager.get_screen('login').ids.drop_item.text, 'OK']
        adicionar = ';'.join(adicionar)
        # salvar observação no arquivo
        with open(os.path.join('dados.txt'),
                  'a') as f:
            f.write(f'\n{adicionar}')

        self.dialog_ok = MDDialog(text="Assinado com sucesso!", radius=[20, 7, 20, 7], )
        self.dialog_ok.open()


class Content(BoxLayout):
    pass


class TelaCadastro(Screen):

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.dialog_atu = None
        self.dialog_add = None
        self.dialog_apg = None
        self.tabela_cadastro = None
        self.lista_cadastro = []
        self.cad_dialog = None

    def cadastro_contas(self):
        self.lista_cadastro.clear()
        conn = sqlite3.connect('contas')
        cursor = conn.cursor()
        cursor.execute('select * from cadastro order by Conta')
        dados = cursor.fetchall()
        for i in dados:
            self.lista_cadastro.append(i)
        conn.close()
        self.tabela_cadastro = MDDataTable(pos_hint={'x': 0.2, 'y': 0.2},
                                           size_hint=(0.3, 0.7),
                                           use_pagination=True, rows_num=30,
                                           background_color_header=get_color_from_hex("#03a9e0"),
                                           check=True,
                                           column_data=[("[color=#ffffff]Conta[/color]", dp(50)),
                                                        ("[color=#ffffff]Usuario[/color]", dp(50)),
                                                        ],
                                           row_data=self.lista_cadastro, elevation=1)

        self.add_widget(self.tabela_cadastro)
        self.tabela_cadastro.bind(on_check_press=self.conta_marcada)

    def conta_marcada(self, instance_table, current_row):
        conn = sqlite3.connect('contas')
        cursor = conn.cursor()
        cursor.execute('select * from cadastro where Conta = ?', (current_row[0], ))
        linha = cursor.fetchone()
        self.ids.conta.text = linha[0]
        self.ids.usuario.text = linha[1]

    def adicionar_conta(self):
        conn = sqlite3.connect('contas')
        cursor = conn.cursor()
        cursor.execute('insert into cadastro (Conta, Usuario) values (?, ?)',
                       (self.ids.conta.text, self.ids.usuario.text))
        conn.commit()
        self.cadastro_contas()
        self.dialog_add = MDDialog(text="Conta Adicionada com sucesso!", radius=[20, 7, 20, 7], )
        self.dialog_add.open()

    def atualizar_conta(self):
        conn = sqlite3.connect('contas')
        cursor = conn.cursor()
        cursor.execute('update cadastro set Usuario = ? where Conta = ?',
                       (self.ids.usuario.text, self.ids.conta.text))
        conn.commit()
        self.cadastro_contas()
        self.dialog_atu = MDDialog(text="Conta atualizada com sucesso!", radius=[20, 7, 20, 7], )
        self.dialog_atu.open()

    def apagar_conta(self):
        conn = sqlite3.connect('contas')
        cursor = conn.cursor()
        cursor.execute('delete from cadastro where Conta = ?',
                       (self.ids.conta.text, ))
        conn.commit()
        self.cadastro_contas()
        self.dialog_apg = MDDialog(text="Conta apagada com sucesso!", radius=[20, 7, 20, 7], )
        self.dialog_apg.open()


class WindowManager(ScreenManager):
    pass


class Conciliacoes(MDApp):
    Window.maximize()
    popupWindow = None

    def meu_popup(self):
        Conciliacoes.popupWindow = MDDialog(text="Gerando relatório...", radius=[20, 7, 20, 7], )
        Conciliacoes.popupWindow.open()

    def build(self):
        pass


Conciliacoes().run()

